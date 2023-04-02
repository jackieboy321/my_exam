import os
import json
#import random
import openpyxl
import numpy as np
import sys

def similarity(a, b):
    a_chars = set(a)
    b_chars = set(b)
    common_chars = a_chars.intersection(b_chars)
    return len(common_chars) / len(a_chars)

def get_similar_questions(sheet, target_question, similarity_threshold=0.8):
    similar_questions = []
    max_rows = sheet.max_row
    for row in range(1, max_rows):
        question = sheet.cell(row=row, column=1).value
        if similarity(target_question, question) > similarity_threshold:
            similar_question = {key: sheet.cell(row=row, column=col).value for col, key in enumerate(['题目', 'A', 'B', 'C', 'D'], start=1)}
            similar_question['答案'] = sheet.cell(row=row, column=sheet.max_column).value
            similar_question['题库序号'] = row
            similar_questions.append(similar_question)
    return similar_questions

def load_workbook(file_name):
    return openpyxl.load_workbook(file_name)


def get_weighted_random_questions(sheet, num_questions, exam_statistics, exam_id, question_source):
    max_rows = sheet.max_row
    question_indices = list(range(1, max_rows))
    
    weights = []
    for index in question_indices:
        key = (question_source, index)
        if key in exam_statistics[exam_id]['incorrect'].keys():
            weights.append(2.0 + exam_statistics[exam_id]['incorrect'][key])
        else:
            weights.append(1.0)
    
    weights = np.array(weights)
    weights /= weights.sum()

    random_questions = np.random.choice(question_indices, size=num_questions, replace=False, p=weights)

    questions = []
    for row in random_questions:
        question = {key: sheet.cell(row=row, column=col).value for col, key in enumerate(['题目', 'A', 'B', 'C', 'D'], start=1)}
        question['答案'] = sheet.cell(row=row, column=sheet.max_column).value
        question['题库序号'] = row
        question['来源'] = question_source
        questions.append(question)

    return questions




def get_user_answer(question_type):
    if question_type <= 150:
        while True:
            user_answer = input("请输入答案: ").strip().upper()
            if user_answer in ['A', 'B', 'C', 'D']:
                return user_answer
            else:
                print("输入有误，请输入A、B、C或D")
    else:
        while True:
            user_answer = input("请输入答案: ").strip().upper()
            if user_answer in ['对', '错']:
                return user_answer
            else:
                print("输入有误，请输入'对'或'错'")


def ask_question(question, question_type):
    print(f"{question_type}. {question['题目']}")
    if all(option is not None for option in (question['A'], question['B'], question['C'], question['D'])):
        print(f"   {question['A']}")
        print(f"   {question['B']}")
        print(f"   {question['C']}")
        print(f"   {question['D']}")

    user_answer = get_user_answer(question_type)
    return user_answer == question['答案'], user_answer, question['题库序号'], question['来源']


def load_exam_statistics(file_name):
    if os.path.exists(file_name):
        with open(file_name, 'r') as f:
            exam_statistics = json.load(f)
        return exam_statistics
    else:
        return {}


def save_exam_statistics(file_name, exam_statistics):
    with open(file_name, 'w') as f:
        json.dump(exam_statistics, f)


def main():
    #file_name = "执照理论考试汇总--监视篇.xlsx"
    if getattr(sys, 'frozen', False):
        application_path = sys._MEIPASS
    else:
        application_path = os.path.dirname(os.path.abspath(__file__))

    file_name = os.path.join(application_path, "执照理论考试汇总--监视篇.xlsx")
    wb = load_workbook(file_name)

    sheet1 = wb['Sheet1']
    sheet2 = wb['Sheet2']

    num_choice_questions = 150
    num_true_false_questions = 50

    exam_statistics_file = 'exam_jianshi.json'
    exam_statistics = load_exam_statistics(exam_statistics_file)

    exam_id = input("请输入您的考试ID：")

    if exam_id not in exam_statistics:
        exam_statistics[exam_id] = {'incorrect': {}}

    choice_questions = get_weighted_random_questions(sheet1, num_choice_questions, exam_statistics, exam_id, "Sheet1")
    true_false_questions = get_weighted_random_questions(sheet2, num_true_false_questions, exam_statistics, exam_id, "Sheet2")


    score = 0
    incorrect_questions = []

    for index, question in enumerate(choice_questions, start=1):
        is_correct, user_answer, question_index, question_source = ask_question(question, index)
        if is_correct:
            score += 0.5
        else:
            incorrect_questions.append((index, question, user_answer, question_index, question_source))
            #incorrect_questions.append((index, question, user_answer, question_index))

    for index, question in enumerate(true_false_questions, start=num_choice_questions + 1):
        is_correct, user_answer, question_index, question_source = ask_question(question, index)
        if is_correct:
            score += 0.5
        else:
            incorrect_questions.append((index, question, user_answer, question_index, question_source))
            #incorrect_questions.append((index, question, user_answer, question_index))

    print(f"\n\033[31m您的分数为：{score}\033[0m")
    input("\n按回车键确认，查看错题（同时错题会保存在当前目录的“错题.txt”中）")

    with open('错题.txt', 'w', encoding='utf-8') as f:
        if incorrect_questions:
            print("\n以下是您的错题：")
            print("\n以下是您的错题：", file=f)
            for index, question, user_answer, question_index, question_source in incorrect_questions:
                print("============================================================================================================================", file=f)
                print(f"{index}. {question['题目']}")
                if all(option is not None for option in (question['A'], question['B'], question['C'], question['D'])):
                    print(f"   {question['A']}\n   {question['B']}\n   {question['C']}\n   {question['D']}")
                #print(f"   {question['A']}\n   {question['B']}\n   {question['C']}\n   {question['D']}")
                print(f"您的答案: {user_answer}")
                print(f"\033[31m正确答案: {question['答案']}\033[0m")
                print()
                print(f"{index}. {question['题目']}", file=f)
                if all(option is not None for option in (question['A'], question['B'], question['C'], question['D'])):
                    print(f"   {question['A']}\n   {question['B']}\n   {question['C']}\n   {question['D']}", file=f)
                #print(f"   {question['A']}\n   {question['B']}\n   {question['C']}\n   {question['D']}", file=f)
                #print(f"您的答案: {user_answer}", file=f)
                print(f"正确答案: {question['答案']}", file=f)
                print("", file=f)
                key = f"{question_source}-{question_index}"
                if key in exam_statistics[exam_id]['incorrect']:
                    exam_statistics[exam_id]['incorrect'][key] += 1
                else:
                    exam_statistics[exam_id]['incorrect'][key] = 1

    with open('错题类似.txt', 'w', encoding='utf-8') as f_similar:
        if incorrect_questions:
            print("\n以下是您的错题及相似题目：", file=f_similar)
            for index, question, user_answer, question_index, question_source in incorrect_questions:
                print("============================================================================================================================", file=f_similar)
                print(f"{index}. {question['题目']}", file=f_similar)
                if all(option is not None for option in (question['A'], question['B'], question['C'], question['D'])):
                    print(f"   {question['A']}\n   {question['B']}\n   {question['C']}\n   {question['D']}", file=f_similar)
                print(f"您的答案: {user_answer}", file=f_similar)
                print(f"正确答案: {question['答案']}", file=f_similar)
                print("", file=f_similar)

                if question_source == "Sheet1":
                    sheet = sheet1
                else:
                    sheet = sheet2

                similar_questions = get_similar_questions(sheet, question['题目'])
                if similar_questions:
                    print("相似题目：", file=f_similar)
                    for similar_question in similar_questions:
                        if similar_question['题库序号'] != question_index:
                            print(f"{similar_question['题目']}", file=f_similar)
                            if all(option is not None for option in (similar_question['A'], similar_question['B'], similar_question['C'], similar_question['D'])):
                                print(f"   {similar_question['A']}\n   {similar_question['B']}\n   {similar_question['C']}\n   {similar_question['D']}", file=f_similar)
                            print(f"正确答案: {similar_question['答案']}", file=f_similar)
                            print("", file=f_similar)
            

    save_exam_statistics(exam_statistics_file, exam_statistics)


if __name__ == "__main__":
    main()
