import os
import json
import random
import openpyxl


def load_workbook(file_name):
    return openpyxl.load_workbook(file_name)

def get_weighted_random_questions(sheet, num_questions, exam_statistics, exam_id):
    max_rows = sheet.max_row
    question_indices = list(range(1, max_rows))
    
    weights = []
    for index in question_indices:
        if index in exam_statistics[exam_id]['correct']:
            weights.append(0.5)
        elif index in exam_statistics[exam_id]['incorrect']:
            weights.append(2.0)
        else:
            weights.append(1.0)
    
    random_questions = random.choices(question_indices, weights=weights, k=num_questions)

    questions = []
    for row in random_questions:
        question = {key: sheet.cell(row=row, column=col).value for col, key in enumerate(['题目', 'A', 'B', 'C', 'D'], start=1)}
        question['答案'] = sheet.cell(row=row, column=sheet.max_column).value
        questions.append(question)

    return questions

def get_user_answer(question_type):
    if question_type <= 150:
        while True:
            user_answer = input("请输入答案: ").strip().upper()
            if user_answer in ['A', 'B', 'C', 'D']:
                return user_answer
            else:
                print("输入有误，请输入A、B、C或D")
    else:
        while True:
            user_answer = input("请输入答案: ").strip().upper()
            if user_answer in ['对', '错']:
                return user_answer
            else:
                print("输入有误，请输入'对'或'错'")

def ask_question(question, question_type):
    print(f"{question_type}. {question['题目']}")
    if 'A' in question and question['A'] is not None:
        print(f"   {question['A']}")
        print(f"   {question['B']}")
        print(f"   {question['C']}")
        print(f"   {question['D']}")

    user_answer = get_user_answer(question_type)
    return user_answer == question['答案'], user_answer

def load_exam_statistics(file_name):
    if os.path.exists(file_name):
        with open(file_name, 'r') as f:
            exam_statistics = json.load(f)

        for exam_id in exam_statistics:
            exam_statistics[exam_id]['correct'] = set(exam_statistics[exam_id]['correct'])
            exam_statistics[exam_id]['incorrect'] = set(exam_statistics[exam_id]['incorrect'])

        return exam_statistics
    else:
        return {}


def save_exam_statistics(file_name, exam_statistics):
    for exam_id in exam_statistics:
        exam_statistics[exam_id]['correct'] = list(exam_statistics[exam_id]['correct'])
        exam_statistics[exam_id]['incorrect'] = list(exam_statistics[exam_id]['incorrect'])

    with open(file_name, 'w') as f:
        json.dump(exam_statistics, f)


def main():
    file_name = "执照理论考试汇总--导航篇.xlsx"
    wb = load_workbook(file_name)

    sheet1 = wb['Sheet1']
    sheet2 = wb['Sheet2']

    num_choice_questions = 150
    num_true_false_questions = 50

    exam_statistics_file = 'exam_statistics.json'
    exam_statistics = load_exam_statistics(exam_statistics_file)

    exam_id = input("请输入您的考试ID：")

    if exam_id not in exam_statistics:
        exam_statistics[exam_id] = {'correct': set(), 'incorrect': set()}

    choice_questions = get_weighted_random_questions(sheet1, num_choice_questions, exam_statistics, exam_id)
    true_false_questions = get_weighted_random_questions(sheet2, num_true_false_questions, exam_statistics, exam_id)

    score = 0
    incorrect_questions = []

    for index, question in enumerate(choice_questions, start=1):
        is_correct, user_answer = ask_question(question, index)
        if is_correct:
            score += 0.5
        else:
            incorrect_questions.append((index, question, user_answer))

    for index, question in enumerate(true_false_questions, start=num_choice_questions + 1):
        is_correct, user_answer = ask_question(question, index)
        if is_correct:
            score += 0.5
        else:
            incorrect_questions.append((index, question, user_answer))

    print(f"\n\033[31m您的分数为：{score}\033[0m")
    input("\n按回车键确认，查看错题（同时错题会保存在当前目录的“错题.txt”中）")

    with open('错题.txt', 'w', encoding='utf-8') as f:
        if incorrect_questions:
            print("\n以下是您的错题：")
            print("\n以下是您的错题：", file=f)
            for index, question, user_answer in incorrect_questions:
                print(f"{index}. {question['题目']}")
                print(f"{question['A'],question['B'],question['C'],question['D']}")
                print(f"您的答案: {user_answer}")
                print(f"\033[31m正确答案: {question['答案']}\033[0m")
                print()
                print(f"{index}. {question['题目']}", file=f)
                print(f"{question['A'],question['B'],question['C'],question['D']}", file=f)
                print(f"您的答案: {user_answer}", file=f)
                print(f"正确答案: {question['答案']}", file=f)
                print("", file=f)

                exam_statistics[exam_id]['incorrect'].add(index)

    for index, question in enumerate(choice_questions, start=1):
        if index not in exam_statistics[exam_id]['incorrect']:
            exam_statistics[exam_id]['correct'].add(index)

    for index, question in enumerate(true_false_questions, start=num_choice_questions + 1):
        if index not in exam_statistics[exam_id]['incorrect']:
            exam_statistics[exam_id]['correct'].add(index)

    save_exam_statistics(exam_statistics_file, exam_statistics)

if __name__ == "__main__":
    main()

